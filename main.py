# 该程序用于将我NAS 网盘中的视频文件压缩到更小
#1. 遍历整个Z盘目录中的所有文件
#2. 判断这个文件是否是视频文件（mov. mp4. )
#3. 如果不是视频文件，continue；如果是视频文件，则判断是否是数字开头的文件，如果不是，说明有问题，exit退出
#4. 读取文件前面4个数字存放到year中，56两个数字存放到month中
#   src = 当前目录   dst = dst\year\month (如果dst不存在，则创建dst先），接下来执行move操作

import os
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook

def compressvideo (srcfile, output):
    name = srcfile.split('\\')[-1]
    outputfile = output
    os.system("ffmpeg -i {} -c:v libx265 -x265-params crf=18:preset=placebo {}".format(srcfile, outputfile))


# Press the green button in the gutter to run the script.
if __name__ == '__main__':

    # 设置初始变量
    srcpathroot = "Y:\\temp"
    dstpathroot = "Y:\\temp"
    videofileext = ('mov', 'mp4', 'avi', 'mkv','flv','mpg','3gp')
    pictureext = ('jpg','png')

    failedfile = []
    movedfilesrc = []
    movedfiledst = []
    movednumber = 0

    #接受用户输入
    print('this is going to compress the video and move it to the desired place')
    # userinput = input('please input your first src and then dst path in strings, if you input the enter directly, '
    #                   'will use the default directors:')
    # if userinput:
    #     srcpathroot, dstpathroot = userinput.split(' ')
    # if dstpathroot == '':
    #     dstpathroot = srcpathroot

    print(f" search src is {srcpathroot}, destination director is {dstpathroot}")
    fileext = videofileext
    for dirpath, dirnames, files in os.walk(srcpathroot, topdown=False):
        print(f'Found directory: {dirpath}')

        for files_name in files:
            suffix = files_name.lower().endswith(tuple(fileext))
            # 得到文件的全路径+文件名
            srcpath = os.path.join(dirpath, files_name)
            # 只有是视频文件才处理
            if suffix:
                year = files_name[0:4]
                month = files_name[4:6]

                if year.isdigit() is False or month.isdigit() is False:
                    failedfile.append(srcpath)
                    continue

                if int(year) not in range(2000,2030) or int(month) not in range(1,13):
                    failedfile.append(srcpath)
                    continue

                # 得到目的地路径
                dstpath = os.path.join(dstpathroot, year, month)
                # 如果该路径不存在，则创建该路径下对应的全部目录
                if not os.path.exists(dstpath):
                    os.makedirs(dstpath)
                # 得到需要移动的文件的全路径
                dstpath = os.path.join(dstpath, files_name)
                # 保存源路径和目的地全路径
                movedfilesrc.append(srcpath)
                movedfiledst.append(dstpath)
                movednumber = movednumber + 1

            else:
                failedfile.append(srcpath)

    # 终于可以开始批量压缩和移动文件
    print(movednumber, " files is going to be compressed")
    compressedfilenumber = 0
    for i in range(0, movednumber):
        if os.path.getsize(movedfilesrc[i]) > 100*1024*1024:
            print(f"compressing the file of {movedfilesrc[i]}")
            compressvideo(movedfilesrc[i], movedfiledst[i])
            compressedfilenumber += 1

    print(movednumber, " files is going to be batched moved")
    for i in range(0, movednumber):
        if dstpathroot != srcpathroot:
            print(f"moving the file from {movedfilesrc[i]}, {movedfiledst[i]}")
            shutil.move(movedfilesrc[i], movedfiledst[i])

    # 结束，保存执行的结果
    print(movednumber, " files were handled")
    print("start to store the result into the result.xlsx")

    wb = load_workbook('result.xlsx')
    ws = wb.active

    for idx in range(1, len(failedfile)+1):
        ws.cell(idx, 1).value = failedfile[idx-1]
    for idx in range(1, len(movedfilesrc)+1):
        ws.cell(idx, 4).value = movedfilesrc[idx-1]
        ws.cell(idx, 6).value = movedfiledst[idx-1]
    wb.save("result.xlsx")

